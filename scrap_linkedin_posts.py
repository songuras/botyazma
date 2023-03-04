from openpyxl import Workbook
from openpyxl   import load_workbook
import os


#excel dosyasını oluşturuyoruz

excel_file = "linkedin_posts.xlsx"

if os.path.exists(excel_file): # eğer dosya varsa
    wb = load_workbook(excel_file)
    ws = wb.active
else: #dosya yoksa
    wb = Workbook()
    ws = wb.active
    ws.append([
        "owner name"
        "owner url"
        "date"
        "text"
        "shared text"
    ])









wb.save(excel_file)